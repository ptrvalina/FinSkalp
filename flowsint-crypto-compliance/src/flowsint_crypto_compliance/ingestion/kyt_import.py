"""Import MetaSleuth / BlockSec / generic KYT wallet exports into label cache."""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from flowsint_types.fiat_crypto import Chain, RegistrySource, SovereignRiskLabel

# Column aliases seen in MetaSleuth, BlockSec, Chainalysis bulk, etc.
_ADDRESS_KEYS = frozenset(
    {"address", "wallet", "wallet_address", "counterparty", "counterparty_address", "from", "to"}
)
_ENTITY_KEYS = frozenset(
    {"entity", "entity_name", "label", "name", "owner", "service", "entity_label"}
)
_CATEGORY_KEYS = frozenset(
    {"category", "tag", "tags", "type", "label_type", "entity_type", "risk_tag"}
)
_RISK_KEYS = frozenset({"risk_score", "risk", "risk_level", "risk_pct", "score"})
_HOPS_KEYS = frozenset({"hops", "hop_count", "hop", "count_hops"})
_AMOUNT_KEYS = frozenset({"amount", "total_received", "value", "volume", "sum"})
_BEHAVIOR_KEYS = frozenset({"behavior", "flow_type", "exposure_type"})
_FOCUS_KEYS = frozenset({"focus_address", "target_address", "screened_address", "root_address"})


def parse_kyt_file(path: Path) -> dict[str, Any]:
    """Parse JSON, JSONL, CSV, or XLSX KYT export."""
    suffix = path.suffix.lower()
    if suffix in {".json"}:
        raw = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(raw, dict) and raw.get("exposure_rows") and not isinstance(raw.get("rows"), list):
            labels = []
            for row in raw.get("labels") or []:
                if isinstance(row, dict):
                    lbl = _row_to_label({_norm_header(k): v for k, v in row.items()}, Chain(str(raw.get("chain", "tron")).lower()))
                    if lbl:
                        labels.append(lbl)
                elif hasattr(row, "model_dump"):
                    labels.append(row)
            return {
                "meta": raw.get("meta") or {},
                "labels": labels,
                "exposure_rows": [_normalize_exposure_row({_norm_header(k): v for k, v in r.items()}) for r in raw["exposure_rows"]],
                "focus_address": raw.get("focus_address"),
                "chain": raw.get("chain", "tron"),
                "row_count": len(raw["exposure_rows"]),
            }
        if isinstance(raw, list):
            return _rows_to_bundle(raw)
        if isinstance(raw, dict) and "rows" in raw:
            return _rows_to_bundle(raw["rows"], meta=raw.get("meta") or {})
        return _rows_to_bundle([raw], meta=raw.get("meta") if isinstance(raw, dict) else {})
    if suffix in {".jsonl", ".ndjson"}:
        rows = []
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line:
                rows.append(json.loads(line))
        return _rows_to_bundle(rows)
    if suffix == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as f:
            return _rows_to_bundle(list(csv.DictReader(f)))
    if suffix in {".xlsx", ".xls"}:
        return _parse_xlsx(path)
    raise ValueError(f"Unsupported KYT file type: {suffix}")


def parse_kyt_bytes(data: bytes, filename: str) -> dict[str, Any]:
    suffix = Path(filename).suffix.lower() or ".json"
    import tempfile

    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(data)
        path = Path(tmp.name)
    try:
        return parse_kyt_file(path)
    finally:
        path.unlink(missing_ok=True)


def _parse_xlsx(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    all_rows: list[dict[str, Any]] = []
    exposure_rows: list[dict[str, Any]] = []
    meta: dict[str, Any] = {"sheets": wb.sheetnames}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue
        headers = [_norm_header(h) for h in header_row if h is not None]
        if not headers:
            continue
        sheet_rows: list[dict[str, Any]] = []
        for values in rows_iter:
            if not values or all(v is None or str(v).strip() == "" for v in values):
                continue
            row = {
                headers[i]: values[i]
                for i in range(min(len(headers), len(values)))
                if headers[i]
            }
            sheet_rows.append(row)
        lower_name = sheet_name.lower()
        if any(k in lower_name for k in ("exposure", "connection", "indirect", "entity")):
            exposure_rows.extend(sheet_rows)
        else:
            all_rows.extend(sheet_rows)
    wb.close()
    bundle = _rows_to_bundle(all_rows or exposure_rows, meta=meta)
    if exposure_rows and all_rows:
        bundle["exposure_rows"] = [_normalize_exposure_row(r) for r in exposure_rows]
    return bundle


def _rows_to_bundle(rows: list[dict[str, Any]], meta: dict[str, Any] | None = None) -> dict[str, Any]:
    labels: list[SovereignRiskLabel] = []
    exposure: list[dict[str, Any]] = []
    focus_address: str | None = None
    chain: Chain | None = None

    for row in rows:
        norm = {_norm_header(k): v for k, v in row.items() if v is not None and str(v).strip() != ""}
        if not norm:
            continue
        focus_address = focus_address or _pick(norm, _FOCUS_KEYS)
        chain = chain or _infer_chain(norm, focus_address)
        exp = _normalize_exposure_row(norm)
        if exp.get("entity_name") or exp.get("category"):
            exposure.append(exp)
        label = _row_to_label(norm, chain or Chain.TRON)
        if label:
            labels.append(label)

    return {
        "meta": meta or {},
        "labels": labels,
        "exposure_rows": exposure,
        "focus_address": focus_address,
        "chain": (chain or Chain.TRON).value,
        "row_count": len(rows),
    }


def _row_to_label(row: dict[str, Any], default_chain: Chain) -> SovereignRiskLabel | None:
    address = _pick(row, _ADDRESS_KEYS)
    if not address:
        return None
    address = str(address).strip()
    if not _looks_like_address(address):
        return None
    entity = _pick(row, _ENTITY_KEYS)
    category = _pick(row, _CATEGORY_KEYS)
    if isinstance(category, str) and "," in category:
        category = category.split(",")[0].strip()
    risk_raw = _pick(row, _RISK_KEYS)
    risk_score = _parse_risk_score(risk_raw)
    sanctioned = _is_sanctioned(row, category, entity)
    if not entity and not category and risk_score is None and not sanctioned:
        return None
    chain = _infer_chain(row, address) or default_chain
    label_id = f"kyt-{chain.value}-{address[:12]}-{entity or category or 'tag'}"
    return SovereignRiskLabel(
        label_id=label_id[:128],
        source=RegistrySource.INTERNAL_OSINT,
        chain=chain,
        address=address,
        entity_name=str(entity) if entity else None,
        category=str(category).lower() if category else None,
        risk_score=risk_score,
        confidence=0.85,
        sanctioned=sanctioned,
        list_reference="kyt_import:metasleuth|blocksec",
    )


def _normalize_exposure_row(row: dict[str, Any]) -> dict[str, Any]:
    entity = _pick(row, _ENTITY_KEYS) or "Unnamed"
    category = (_pick(row, _CATEGORY_KEYS) or "unknown").lower()
    risk_pct = _parse_risk_pct(_pick(row, _RISK_KEYS))
    hops = _parse_int(_pick(row, _HOPS_KEYS))
    amount = _parse_float(_pick(row, _AMOUNT_KEYS))
    behavior = (_pick(row, _BEHAVIOR_KEYS) or "indirect").lower()
    tier = _category_tier(category, risk_pct)
    return {
        "entity_name": str(entity),
        "category": category,
        "risk_pct": risk_pct,
        "risk_tier": tier,
        "hops": hops,
        "amount": amount,
        "behavior": behavior,
        "address": _pick(row, _ADDRESS_KEYS),
    }


def import_kyt_bundle(cache: Any, bundle: dict[str, Any]) -> dict[str, int]:
    """Upsert labels from bundle into LabelCache and EntityLabelStore."""
    from flowsint_crypto_compliance.attribution.entity_label_store import get_entity_label_store
    from flowsint_crypto_compliance.attribution.types import TIER_CONFIRMED_IMPORT, EntityLabel

    labels = bundle.get("labels") or []
    count = cache.bulk_upsert(labels)
    el_store = get_entity_label_store()
    for lbl in labels:
        el_store.upsert(
            EntityLabel(
                address=lbl.address,
                chain=lbl.chain.value,
                label=lbl.entity_name or lbl.category or "imported",
                category=(lbl.category or "other").lower(),
                confidence=lbl.confidence,
                source="kyt_import",
                tier=TIER_CONFIRMED_IMPORT,
                risk_score=lbl.risk_score or 30.0,
                sanctioned=lbl.sanctioned,
                evidence="kyt_import",
            ),
            force=True,
        )
    for row in bundle.get("exposure_rows") or []:
        addr = row.get("address")
        if addr and row.get("entity_name"):
            el_store.upsert(
                EntityLabel(
                    address=addr,
                    chain=bundle.get("chain", "tron"),
                    label=str(row.get("entity_name")),
                    category=str(row.get("category") or "other"),
                    confidence=0.9,
                    source="kyt_import",
                    tier=TIER_CONFIRMED_IMPORT,
                    risk_score=float(row.get("risk_pct") or 20),
                    evidence="kyt_import:exposure",
                ),
                force=True,
            )
    return {
        "labels_imported": count,
        "exposure_rows": len(bundle.get("exposure_rows") or []),
        "total_in_cache": cache.count(),
        "entity_labels": el_store.count(),
    }


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s\-]+", "_", text)
    return text


def _pick(row: dict[str, Any], keys: frozenset[str]) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def _looks_like_address(value: str) -> bool:
    if re.fullmatch(r"0x[a-fA-F0-9]{40}", value):
        return True
    if re.fullmatch(r"T[1-9A-HJ-NP-Za-km-z]{33}", value):
        return True
    if re.fullmatch(r"([13][a-km-zA-HJ-NP-Z1-9]{25,34}|bc1[a-z0-9]{39,59})", value):
        return True
    return False


def _infer_chain(row: dict[str, Any], address: str | None) -> Chain | None:
    raw = row.get("chain") or row.get("network") or row.get("blockchain")
    if raw:
        try:
            return Chain(str(raw).lower())
        except ValueError:
            pass
    if address:
        if address.startswith("0x"):
            return Chain.ETH
        if address.startswith("T") and len(address) == 34:
            return Chain.TRON
        if address.startswith(("bc1", "1", "3")):
            return Chain.BTC
    return None


def _parse_risk_score(value: Any) -> float | None:
    if value is None or value == "":
        return None
    text = str(value).strip().lower()
    if text in {"critical", "severe", "high"}:
        return {"critical": 95.0, "severe": 90.0, "high": 75.0}[text]
    if text in {"medium", "moderate"}:
        return 50.0
    if text in {"low"}:
        return 15.0
    try:
        num = float(text.replace("%", ""))
        return num if num <= 100 else num / 100.0
    except ValueError:
        return None


def _parse_risk_pct(value: Any) -> float:
    score = _parse_risk_score(value)
    if score is None:
        return 0.0
    return score if score <= 100 else score


def _parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value)))
    except ValueError:
        return None


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return None


def _is_sanctioned(row: dict[str, Any], category: Any, entity: Any) -> bool:
    text = " ".join(str(v).lower() for v in [category, entity, row.get("sanctioned")] if v)
    if str(row.get("sanctioned", "")).lower() in {"1", "true", "yes"}:
        return True
    return any(k in text for k in ("sanction", "ofac", "terror", "extremist"))


def _category_tier(category: str, risk_pct: float) -> str:
    cat = category.lower()
    if any(k in cat for k in ("sanction", "mixer", "scam", "ransom", "terror", "illicit", "risky_exchange")):
        return "severe"
    if any(k in cat for k in ("gambling", "darknet", "fraud", "illegal")):
        return "high"
    if risk_pct >= 70:
        return "high"
    if risk_pct >= 40:
        return "moderate"
    return "low"
