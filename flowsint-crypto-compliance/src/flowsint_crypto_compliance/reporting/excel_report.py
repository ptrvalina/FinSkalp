"""Excel export for regulator compliance reports."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


def render_regulator_xlsx(report: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    bold = Font(bold=True)

    ws["A1"] = "Кейс"
    ws["B1"] = report.get("case_ref", "")
    ws["A2"] = "Сценарий"
    ws["B2"] = report.get("scenario_title_ru", "")
    ws["A3"] = "Индекс риска"
    ws["B3"] = report.get("illegal_flow_score", "")
    ws["A4"] = "Уровень"
    ws["B4"] = report.get("risk_level", "")
    ws["A5"] = "Резюме"
    ws["B5"] = report.get("executive_summary_ru", "")
    for cell in ("A1", "A2", "A3", "A4", "A5"):
        ws[cell].font = bold

    ws.append([])
    ws.append(["Индикаторы"])
    ws["A7"].font = bold
    ws.append(["Severity", "Code", "Title", "Confidence"])
    for finding in report.get("findings", []):
        ws.append(
            [
                finding.get("severity"),
                finding.get("code"),
                finding.get("title_ru"),
                finding.get("confidence"),
            ]
        )

    metrics = report.get("metrics") or {}
    xgb = (metrics.get("risk_scoring") or {}).get("xgboost") or {}
    ws.append([])
    ws.append(["XGBoost"])
    ws[f"A{ws.max_row}"].font = bold
    ws.append(["Model", xgb.get("model_version", "")])
    ws.append(["Heuristic", metrics.get("risk_scoring", {}).get("heuristic_score", "")])
    ws.append(["Model score", xgb.get("model_score", "")])
    ws.append(["Blended", xgb.get("blended_score", "")])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
